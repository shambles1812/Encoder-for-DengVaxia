from selenium import webdriver
from selenium.webdriver.common.keys import Keys 
from openpyxl import load_workbook, Workbook
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
import string
import time

#load Workbook
wb = load_workbook('Database2.xlsx')
alphabet = list(string.ascii_uppercase)
#alphabet = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']

#cell colors
greenFill = PatternFill(start_color='2ecc71',
                   end_color='2ecc71',
                   fill_type='solid')
redFill = PatternFill(start_color='e74c3c',
                   end_color='e74c3c',
                   fill_type='solid')



#load resources for a specific worksheet and user
class get_worksheet_data:
    def __init__(self,worksheet_no,user_number):
        #choose worksheet
        self.worksheet_no = worksheet_no
        self.user_number = user_number
        sheet = wb.worksheets[worksheet_no-1]
        self.sheet = sheet
        wb.active = sheet
        #get the number of people from page
        profiles_to_encode = sheet.max_row - 1 
        self.profiles_to_encode = profiles_to_encode
        #get the number of data from page
        data_to_encode = sheet.max_column 
        self.data_to_encode = data_to_encode
        #profile 1 profile 2 profile 3 etc..
        profile_list = list(range(1,profiles_to_encode+1))
        self.profile_list = profile_list
        data_list = list(range(1,data_to_encode+1))
        self.data_list = data_list
        profile_users = []
        i= 0
        #create list of 
        for profile in profile_list:
            profile_data = []
            for letter in alphabet:
                test = [str(letter+str(profile+1))]
                i = i + 1
                profile_data.extend(test)
                if i % len(alphabet) == 0:
                    profile_users.append(profile_data)
        self.profile_users = profile_users

    def get_all_data(self):
        for data in self.data_list:
            print(str(self.user_number-1) + str(data-1))
            if self.sheet[self.profile_users[self.user_number-1][data-1]].value:
                print(self.sheet[self.profile_users[self.user_number-1][data-1]].value) 
    def get_specific_data(self,index_no):
            if self.sheet[self.profile_users[self.user_number-1][index_no-1]].value:
                return self.sheet[self.profile_users[self.user_number-1][index_no-1]].value
    def set_to_ok(self,index_no):
            self.sheet[self.profile_users[self.user_number-1][index_no-1]].fill = greenFill
           # wb.save('Stylized.xlsx')
    def set_to_not_ok(self,index_no):
            self.sheet[self.profile_users[self.user_number-1][index_no-1]].fill = redFill
           # wb.save('Stylized.xlsx')



        
pageData1 = get_worksheet_data(1,1)
print (get_worksheet_data(1,1).get_specific_data(1))

#main loop 
# login user -> navigate to encoding page -> encode[ input on fields -> change cell color -> click save] 
for profile in pageData1.profile_list:
    
    #login
    browser = webdriver.Chrome('C:\\Users\\Cj\\Desktop\\Temp\\chromedriver')
    browser.get('http://122.54.82.254:65428/')
    user_field = browser.find_element_by_id('UserName')
    pass_field = browser.find_element_by_id('Password')
    login = browser.find_element_by_xpath("//input[@value='Log in']")

    login.click()

    #navigate to vaccination entry page
    create_new_profile = browser.find_element_by_xpath("//button[@data-toggle='modal']")
    time.sleep(2)
    create_new_profile.click()
    add_vacinee_profile = browser.find_element_by_xpath("//a[@href='/VaccinneeProfile/Add']")
    time.sleep(2)
    add_vacinee_profile.click()
    time.sleep(2)

    # paths on page 1
    barangay_option = browser.find_element_by_id('s2id_PSGC_BarangayId')
    id_field = browser.find_element_by_id('VaccinationCardId')
    date_interviewed_field = browser.find_element_by_id('DateInterviewed')
    relation_to_vaccined = browser.find_element_by_id('RelasyonSaBakunado')
    last_name_field = browser.find_element_by_id('LastName')
    first_name_field = browser.find_element_by_id('FirstName')
    middle_inintial_field = browser.find_element_by_id('MiddleName')
    extension_name_field = browser.find_element_by_id('ExtensionName')
    relation_to_household = browser.find_element_by_id('RelationToHouseholdHead')
    respondent = browser.find_element_by_id('Respondent')
    contact_no_field = browser.find_element_by_id('ContactNumber')
    female_selection = browser.find_element_by_xpath("//input[@value='Female']")
    male_selection = browser.find_element_by_xpath("//input[@value='Male']")
    age_field = browser.find_element_by_id('Age')
    birthplace_field = browser.find_element_by_id('BirthPlace')
    years_at_curr = browser.find_element_by_id('YearsInAddress')
    province_option = browser.find_element_by_id('s2id_PSGC_ProvinceId')
    city_option = browser.find_element_by_id('s2id_PSGC_CityMunicipalityId')
    
    street_field = browser.find_element_by_id('Street')
    house_no_field = browser.find_element_by_id('LotNo')
    birthdate_field = browser.find_element_by_id('BirthDate')
    roman_catholic_selection = browser.find_element_by_xpath("//input[@value='Roman Catholic']")
    christian_selection = browser.find_element_by_xpath("//input[@value='Christian']")
    inc_selection = browser.find_element_by_xpath("//input[@value='INC']")
    islam_selection = browser.find_element_by_xpath("//input[@value='Islam']")
    jehovah_selection = browser.find_element_by_xpath("//input[@value='Jehovah']")
    other_selection = browser.find_element_by_xpath("//input[@value='Others']")
    otherReligion_field = browser.find_element_by_xpath("//input[@class='form-control text-box single-line']")
    forgotBirthdate_selection = browser.find_element_by_id('called')
    province_field = browser.find_element_by_xpath('//input[@class="select2-input"]')
    barangay_option = browser.find_element_by_id('s2id_PSGC_BarangayId')# change to something more efficient

    def id_input(profile_no):
        if get_worksheet_data(1,profile_no).get_specific_data(1):
            id_field.send_keys(get_worksheet_data(1,profile_no).get_specific_data(1))
            get_worksheet_data(1,profile_no).set_to_ok(1)
        else:
            get_worksheet_data(1,profile_no).set_to_not_ok(1)
    def date_interviewed_input(profile_no):
        if get_worksheet_data(1,profile_no).get_specific_data(2):
            date_interviewed_field.send_keys(get_worksheet_data(1,profile_no).get_specific_data(2).strftime("%m/%d/%Y"))
            get_worksheet_data(1,profile_no).set_to_ok(2)
        else:
            get_worksheet_data(1,profile_no).set_to_not_ok(2)
    def relation_to_vaccined_input(profile_no):
        if get_worksheet_data(1,profile_no).get_specific_data(3):
            relation_to_vaccined.send_keys(get_worksheet_data(1,profile_no).get_specific_data(3))
            get_worksheet_data(1,profile_no).set_to_ok(3)
        else:
            get_worksheet_data(1,profile_no).set_to_not_ok(3)
    def last_name_field_input(profile_no):
        if get_worksheet_data(1,profile_no).get_specific_data(4):
            last_name_field.send_keys(get_worksheet_data(1,profile_no).get_specific_data(4))
            get_worksheet_data(1,profile_no).set_to_ok(4)
        else:
            get_worksheet_data(1,profile_no).set_to_not_ok(4)
    def first_name_field_input(profile_no):
        if get_worksheet_data(1,profile_no).get_specific_data(5):
            first_name_field.send_keys(get_worksheet_data(1,profile_no).get_specific_data(5))
            get_worksheet_data(1,profile_no).set_to_ok(5)
        else:
            get_worksheet_data(1,profile_no).set_to_not_ok(5)
    def middle_inintial_field_input(profile_no):
        if get_worksheet_data(1,profile_no).get_specific_data(6):
            middle_inintial_field.send_keys(get_worksheet_data(1,profile_no).get_specific_data(6))
            get_worksheet_data(1,profile_no).set_to_ok(6)
        else:
            get_worksheet_data(1,profile_no).set_to_not_ok(6)
    def extension_name_field_input(profile_no):
        if get_worksheet_data(1,profile_no).get_specific_data(7):
            extension_name_field.send_keys(get_worksheet_data(1,profile_no).get_specific_data(7))
            get_worksheet_data(1,profile_no).set_to_ok(7)
        else:
            get_worksheet_data(1,profile_no).set_to_not_ok(7)
    def relation_to_household_input(profile_no):
        if get_worksheet_data(1,profile_no).get_specific_data(8):
            first_name_field.send_keys(get_worksheet_data(1,profile_no).get_specific_data(8))
            get_worksheet_data(1,profile_no).set_to_ok(8)
        else:
            get_worksheet_data(1,profile_no).set_to_not_ok(8)
    def respondent_input(profile_no):
        if get_worksheet_data(1,profile_no).get_specific_data(9):
            respondent.send_keys(get_worksheet_data(1,profile_no).get_specific_data(9))
            get_worksheet_data(1,profile_no).set_to_ok(9)
        else:
            get_worksheet_data(1,profile_no).set_to_not_ok(9)
    def contact_no_field_input(profile_no):
        if get_worksheet_data(1,profile_no).get_specific_data(10):
            contact_no_field.send_keys(("0")+str(get_worksheet_data(1,profile).get_specific_data(10)))
            get_worksheet_data(1,profile_no).set_to_ok(10)
        else:
            get_worksheet_data(1,profile_no).set_to_not_ok(10)
    def province_input(profile_no):
        if get_worksheet_data(1,profile_no).get_specific_data(11):
            province_option.click()
            time.sleep(1)
            province_field.send_keys(get_worksheet_data(1,profile_no).get_specific_data(11))
            selector_button = browser.find_element_by_xpath('//li[@class="select2-results-dept-0 select2-result select2-result-selectable select2-highlighted"]')
            selector_button.click()
            get_worksheet_data(1,profile_no).set_to_ok(11)
        else:
            get_worksheet_data(1,profile_no).set_to_not_ok(11)
    def city_input(profile_no):
        if get_worksheet_data(1,profile_no).get_specific_data(12):
            city_option.click()
            city_field = browser.find_element_by_id('s2id_autogen2_search')
            city_field.send_keys(get_worksheet_data(1,profile_no).get_specific_data(12))
            selector_button = browser.find_element_by_xpath('//li[@class="select2-results-dept-0 select2-result select2-result-selectable select2-highlighted"]')
            selector_button.click()
            get_worksheet_data(1,profile_no).set_to_ok(12)
        else:
            get_worksheet_data(1,profile_no).set_to_not_ok(12)
    def barangay_input(profile_no):
        if get_worksheet_data(1,profile_no).get_specific_data(13):
            barangay_option.click()
            barangay_field = browser.find_element_by_id('s2id_autogen3_search')
            barangay_field.send_keys(get_worksheet_data(1,profile_no).get_specific_data(13))
            selector_button = browser.find_element_by_xpath('//li[@class="select2-results-dept-0 select2-result select2-result-selectable select2-highlighted"]')
            selector_button.click()
            get_worksheet_data(1,profile_no).set_to_ok(13)
        else:
            get_worksheet_data(1,profile_no).set_to_not_ok(13)

#test input
    #input to paths
    
    #id_field.send_keys(get_worksheet_data(1,profile).get_specific_data(1))
    id_input(profile)
    #date_interviewed_field.send_keys((get_worksheet_data(1,profile).get_specific_data(2)).strftime("%m/%d/%Y"))
    date_interviewed_input(profile)
    #relation_to_vaccined.send_keys(get_worksheet_data(1,profile).get_specific_data(3))
    relation_to_vaccined_input(profile)
    #last_name_field.send_keys(get_worksheet_data(1,profile).get_specific_data(4))
    last_name_field_input(profile)
    #first_name_field.send_keys(get_worksheet_data(1,profile).get_specific_data(5))
    first_name_field_input(profile)
    #middle_inintial_field.send_keys(get_worksheet_data(1,profile).get_specific_data(6))
    middle_inintial_field_input(profile)
    #extension_name_field.send_keys() #add keys later
    extension_name_field_input(profile)
    #relation_to_household.send_keys(get_worksheet_data(1,profile).get_specific_data(7))
    relation_to_household_input(profile)
    #respondent.send_keys(get_worksheet_data(1,profile).get_specific_data(8))
    respondent_input(profile)
    #contact_no_field.send_keys(("0")+str(get_worksheet_data(1,profile).get_specific_data(9)))
    contact_no_field_input(profile)

    #choose from dropdown
    #province_option.click()
    province_input(profile)
    #time.sleep(1)
    #province_field.send_keys(get_worksheet_data(1,1).get_specific_data(10))
    #selector_button = browser.find_element_by_xpath('//li[@class="select2-results-dept-0 select2-result select2-result-selectable select2-highlighted"]')
    #selector_button.click()

    #city_option.click()
    #city_field = browser.find_element_by_id('s2id_autogen2_search')

    #city_field.send_keys(get_worksheet_data(1,1).get_specific_data(11))
    #selector_button = browser.find_element_by_xpath('//li[@class="select2-results-dept-0 select2-result select2-result-selectable select2-highlighted"]')
    #selector_button.click()

    #barangay_option.click()
   # barangay_field = browser.find_element_by_id('s2id_autogen3_search')
    #barangay_field.send_keys(get_worksheet_data(1,1).get_specific_data(12))
    #selector_button = browser.find_element_by_xpath('//li[@class="select2-results-dept-0 select2-result select2-result-selectable select2-highlighted"]')
    #selector_button.click()

    street_field.send_keys(get_worksheet_data(1,1).get_specific_data(13))
    house_no_field.send_keys(get_worksheet_data(1,1).get_specific_data(14))
    #sex selection
    if get_worksheet_data(1,1).get_specific_data(15) == 'MALE':
        male_selection.click()
    else:
        female_selection.click()
    age_field.send_keys(get_worksheet_data(1,1).get_specific_data(16))
    #religion 
    if get_worksheet_data(1,1).get_specific_data(17) == "Roman Catholic":
        roman_catholic_selection.click()
    elif get_worksheet_data(1,1).get_specific_data(17) == "Christian":
        christian_selection.click()
    elif get_worksheet_data(1,1).get_specific_data(17) == "INC" :
        inc_selection.click()
    elif get_worksheet_data(1,1).get_specific_data(17) == "Islam" :
        islam_selection.click()
    elif get_worksheet_data(1,1).get_specific_data(17) == "Jehovah" :
        jehovah_selection.click()
    else: 
            other_selection.click()
            otherReligion_field.send_keys(get_worksheet_data(1,1).get_specific_data(17))

    birthdate_field.send_keys((get_worksheet_data(1,1).get_specific_data(18)).strftime("%m/%d/%Y"))
    birthplace_field.send_keys(get_worksheet_data(1,1).get_specific_data(19))
    years_at_curr.send_keys(get_worksheet_data(1,1).get_specific_data(20))
    