from selenium import webdriver
from selenium.webdriver.common.keys import Keys 
from openpyxl import load_workbook, Workbook
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
import string
import time

#load Workbook
wb = load_workbook('Database1.xlsx')
sheet = wb.worksheets[0]
wb.active = sheet
top_left_cell = sheet['B2']
#cell colors
greenFill = PatternFill(start_color='2ecc71',
                   end_color='2ecc71',
                   fill_type='solid')
redFill = PatternFill(start_color='e74c3c',
                   end_color='e74c3c',
                   fill_type='solid')
top_left_cell.fill = greenFill
wb.save("Stylized.xlsx")