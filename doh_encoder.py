from selenium import webdriver
from selenium.webdriver.common.keys import Keys 
from openpyxl import load_workbook, Workbook
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
import string
import time

#load Workbook
wb = load_workbook('Database3.xlsx')
alphabet = list(string.ascii_uppercase)
#cell colors
greenFill = PatternFill(start_color='2ecc71',
                   end_color='2ecc71',
                   fill_type='solid')
redFill = PatternFill(start_color='e74c3c',
                   end_color='e74c3c',
                   fill_type='solid')
#load resources for a specific worksheet and user
class get_worksheet_data:
    def __init__(self,worksheet_no,user_number):
        #choose worksheet
        self.worksheet_no = worksheet_no
        self.user_number = user_number
        sheet = wb.worksheets[worksheet_no-1]
        self.sheet = sheet
        wb.active = sheet
        #get the number of people from page
        profiles_to_encode = sheet.max_row - 1 
        self.profiles_to_encode = profiles_to_encode
        #get the number of data from page
        data_to_encode = sheet.max_column 
        self.data_to_encode = data_to_encode
        #profile 1 profile 2 profile 3 etc..
        profile_list = list(range(1,profiles_to_encode+1))
        self.profile_list = profile_list
        data_list = list(range(1,data_to_encode+1))
        self.data_list = data_list
        profile_users = []
        i= 0
        #create list of 
        for profile in profile_list:
            profile_data = []
            for letter in alphabet:
                test = [str(letter+str(profile+1))]
                i = i + 1
                profile_data.extend(test)
                if i % len(alphabet) == 0:
                    profile_users.append(profile_data)
        self.profile_users = profile_users
    def get_all_data(self):
        for data in self.data_list:
            print(str(self.user_number-1) + str(data-1))
            if self.sheet[self.profile_users[self.user_number-1][data-1]].value:
                print(self.sheet[self.profile_users[self.user_number-1][data-1]].value) 
    def get_specific_data(self,index_no):
            if self.sheet[self.profile_users[self.user_number-1][index_no-1]].value:
                return self.sheet[self.profile_users[self.user_number-1][index_no-1]].value
    def set_to_ok(self,index_no):
            self.sheet[self.profile_users[self.user_number-1][index_no-1]].fill = greenFill
           # wb.save('Stylized.xlsx')
    def set_to_not_ok(self,index_no):
            self.sheet[self.profile_users[self.user_number-1][index_no-1]].fill = redFill
           # wb.save('Stylized.xlsx')
pageData1 = get_worksheet_data(1,1)
print (get_worksheet_data(1,1).get_specific_data(1))
#main loop 
# login user -> navigate to encoding page -> encode[ input on fields -> change cell color -> click save] 
for profile in pageData1.profile_list:
    #login
    browser = webdriver.Chrome('C:\\Users\\Cj\\Desktop\\Temp\\chromedriver')
    browser.get('http://122.54.82.254:65428/')
    user_field = browser.find_element_by_id('UserName')
    pass_field = browser.find_element_by_id('Password')
    login = browser.find_element_by_xpath("//input[@value='Log in']")
    user_input = "NATE155"
    pass_input = "123456"
    user_field.send_keys(user_input)
    pass_field.send_keys(pass_input)
    login.click()
    #navigate to vaccination entry page
    create_new_profile = browser.find_element_by_xpath("//button[@data-toggle='modal']")
    time.sleep(2)
    create_new_profile.click()
    add_vacinee_profile = browser.find_element_by_xpath("//a[@href='/VaccinneeProfile/Add']")
    time.sleep(2)
    add_vacinee_profile.click()
    time.sleep(2)
    # first field paths
    id_field = browser.find_element_by_id('VaccinationCardId')
    #lvl_of_educ_field = browser.find_element_by_id('LevelOfEducation')
# General Information functions
    #checks if the first field has input
    def not_filled(first_field):
        first_field.click()
        current_element = browser.switch_to_active_element()
        if len(current_element.get_attribute("value")) == 0:
            return True
        else:
            return False
    def id_input(profile_no):
        if get_worksheet_data(1,profile_no).get_specific_data(1):
            id_field.send_keys(get_worksheet_data(1,profile_no).get_specific_data(1))
            get_worksheet_data(1,profile_no).set_to_ok(1)
        else:
            get_worksheet_data(1,profile_no).set_to_not_ok(1)
    def press_tab():
        current_element = browser.switch_to_active_element()
        current_element.send_keys(Keys.TAB)
    def press_enter():
        current_element = browser.switch_to_active_element()
        current_element.send_keys(Keys.ENTER)
    def press_right_arrow_key():
        current_element = browser.switch_to_active_element()
        current_element.send_keys(Keys.ARROW_RIGHT)
    def press_left_arrow_key():
        current_element = browser.switch_to_active_element()
        current_element.send_keys(Keys.ARROW_LEFT)
    def click_element():
        current_element = browser.switch_to_active_element()
        current_element.click()
#new method for encoding data on General Info Page
    def input_data_on_gen_info(profile):
        #select first element
        id_input(profile)
        press_tab()
        current_element = browser.switch_to_active_element()
        if get_worksheet_data(1,profile).get_specific_data(2):
            current_element.send_keys(get_worksheet_data(1,profile).get_specific_data(2).strftime("%m/%d/%Y"))
            get_worksheet_data(1,profile).set_to_ok(2)
        else: 
            get_worksheet_data(1,profile).get_specific_data(2).set_to_not_ok
        press_tab()
        for number in list(range(3,10)):
            current_element = browser.switch_to_active_element()
            if get_worksheet_data(1,profile).get_specific_data(number):
                current_element.send_keys(get_worksheet_data(1,profile).get_specific_data(number))
                get_worksheet_data(1,profile).set_to_ok(number)
            else: 
                get_worksheet_data(1,profile).set_to_not_ok(number)
            press_tab()
            current_element = browser.switch_to_active_element()
        if get_worksheet_data(1,profile).get_specific_data(10):
            current_element = browser.switch_to_active_element()    
            current_element.send_keys("0"+ str(get_worksheet_data(1,profile).get_specific_data(10)))
            get_worksheet_data(1,profile).set_to_ok(10)
        else: 
            current_element = browser.switch_to_active_element()
            get_worksheet_data(1,profile).set_to_not_ok(10)
        press_tab()
        for number in list(range(11,14)):
            current_element = browser.switch_to_active_element()
            if get_worksheet_data(1,profile).get_specific_data(number):
                press_enter()
                current_element = browser.switch_to_active_element()
                current_element.send_keys(get_worksheet_data(1,profile).get_specific_data(number))
                press_enter()
                current_element = browser.switch_to_active_element()
                get_worksheet_data(1,profile).set_to_ok(number)
            else:
                get_worksheet_data(1,profile).set_to_not_ok(number)
            press_tab()
        for number in list(range(14,16)):
            current_element = browser.switch_to_active_element()
            if get_worksheet_data(1,profile).get_specific_data(number):
                current_element.send_keys(get_worksheet_data(1,profile).get_specific_data(number))
                get_worksheet_data(1,profile).set_to_ok(number)
            else: 
                get_worksheet_data(1,profile).set_to_not_ok(number)
            press_tab()
            current_element = browser.switch_to_active_element()      
        if get_worksheet_data(1,profile).get_specific_data(16) == "MALE":
            press_right_arrow_key()
            current_element = browser.switch_to_active_element()
            press_left_arrow_key()
            current_element = browser.switch_to_active_element()
            get_worksheet_data(1,profile).set_to_ok(16)
        elif get_worksheet_data(1,profile).get_specific_data(16) == "FEMALE":
            press_right_arrow_key()
            current_element = browser.switch_to_active_element()
            press_right_arrow_key()
            current_element = browser.switch_to_active_element()
            get_worksheet_data(1,profile).set_to_ok(16)
        else:
            get_worksheet_data(1,profile).set_to_not_ok(16)
        press_tab()
        current_element = browser.switch_to_active_element()
        if get_worksheet_data(1,profile).get_specific_data(17):
            current_element.send_keys(get_worksheet_data(1,profile).get_specific_data(17))
            get_worksheet_data(1,profile).set_to_ok(17)
        else: 
            get_worksheet_data(1,profile).set_to_not_ok(number)
        press_tab()
        current_element = browser.switch_to_active_element()
        #set religion
        if get_worksheet_data(1,profile).get_specific_data(18):
            if get_worksheet_data(1,profile).get_specific_data(18) == "ROMAN CATHOLIC":
                press_right_arrow_key()
                current_element = browser.switch_to_active_element()
                press_left_arrow_key()
                current_element = browser.switch_to_active_element()
                get_worksheet_data(1,profile).set_to_ok(18)
                press_tab()
                current_element = browser.switch_to_active_element()
            elif get_worksheet_data(1,profile).get_specific_data(18) == "CHRISTIAN":
                press_right_arrow_key()
                current_element = browser.switch_to_active_element()
                get_worksheet_data(1,profile).set_to_ok(18)
                press_tab()
                current_element = browser.switch_to_active_element()
            elif get_worksheet_data(1,profile).get_specific_data(18) == "INC" :
                for number in list(range(1,3)):
                    press_right_arrow_key()
                    current_element = browser.switch_to_active_element() 
                get_worksheet_data(1,profile).set_to_ok(18)
                press_tab()
                current_element = browser.switch_to_active_element()
            elif get_worksheet_data(1,profile).get_specific_data(18) == "ISLAM" :
                for number in list(range(1,4)):
                    press_right_arrow_key()
                    current_element = browser.switch_to_active_element() 
                get_worksheet_data(1,profile).set_to_ok(18)
                press_tab()
                current_element = browser.switch_to_active_element()
            elif get_worksheet_data(1,profile).get_specific_data(18) == "JEHOVAH" :
                for number in list(range(1,5)):
                    press_right_arrow_key()
                    current_element = browser.switch_to_active_element() 
                get_worksheet_data(1,profile).set_to_ok(18)
                press_tab()
                current_element = browser.switch_to_active_element()
            else: 
                for number in list(range(1,6)):
                    press_right_arrow_key()
                    current_element = browser.switch_to_active_element() 
                press_tab()
                current_element = browser.switch_to_active_element()
                current_element.send_keys(get_worksheet_data(1,1).get_specific_data(18))
                get_worksheet_data(1,profile).set_to_ok(18)
        else:
            get_worksheet_data(1,profile).set_to_not_ok(18)
        press_tab()
        if get_worksheet_data(1,profile).get_specific_data(19):
            current_element = browser.switch_to_active_element()
            current_element.send_keys(get_worksheet_data(1,profile).get_specific_data(19).strftime("%m/%d/%Y"))
            get_worksheet_data(1,profile).set_to_ok(19)
            press_tab()
            current_element = browser.switch_to_active_element()
        else: 
            get_worksheet_data(1,profile).set_to_not_ok(19)
            current_element = browser.switch_to_active_element()
            press_tab()
            current_element = browser.switch_to_active_element()
            click_element()
        press_tab()
        for number in list(range(20,22)):
            current_element = browser.switch_to_active_element()
            if get_worksheet_data(1,profile).get_specific_data(number):
                current_element.send_keys(get_worksheet_data(1,profile).get_specific_data(number))
                get_worksheet_data(1,profile).set_to_ok(number)
            else: 
                get_worksheet_data(1,profile).set_to_not_ok(number)
            press_tab() 
        wb.save('Stylized3.xlsx')
        input('wait')
    functions_list = ['input_data_on_gen_info(profile)','id_input']
    #input_to__general_information(profile)
   # make a new for loop that verifies 14 times and executes the input command for each page every loop
    
   #idea is check if the first field has value and if yes skip that page if no fill that page with the right info
    if not_filled(id_field) == True:
        eval(functions_list[0])
    else:

        current_element = browser.switch_to_active_element()
        current_element.send_keys(Keys.ENTER)
    input('wait')